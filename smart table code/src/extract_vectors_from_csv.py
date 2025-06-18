import numpy as np
import pandas as pd
import customtkinter as ctk
import os
from natsort import natsorted
import openpyxl
import matplotlib
matplotlib.use('Agg')  # Use Agg backend for non-GUI environments
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
import io

def calculate_path_length(points):
    """
    Calculate total path length from a sequence of points.
    
    Args:
        points: numpy array of shape (n,2) containing x,y coordinates
        
    Returns:
        float: Total path length
    """
    if points is None or len(points) < 2:
        return 0.0
        
    points = np.array(points)
    # Calculate differences between consecutive points
    diffs = np.diff(points, axis=0)
    # Sum up the Euclidean distances
    return np.sum(np.sqrt(np.sum(diffs**2, axis=1)))

def compute_vectors(csv_path):
    """Compute vectors from CSV file with x,y coordinates and timestamps."""
    try:
        # Read CSV with named columns
        df = pd.read_csv(csv_path, header=None, 
                        names=['x', 'y', 'datetime', 'rep'])
        
        # Convert coordinates to numeric, ensuring 2D array structure
        coords = df[['x', 'y']].values
        
        # Calculate vectors between consecutive points
        vectors = np.zeros((len(coords)-1, 4))  # [x1,y1,dx,dy]
        for i in range(len(coords)-1):
            vectors[i, 0:2] = coords[i]      # Starting point (x1,y1)
            vectors[i, 2:4] = coords[i+1] - coords[i]  # Vector (dx,dy)
            
        return vectors, coords, df['datetime'].tolist(), df['rep'].tolist()
        
    except Exception as e:
        print(f"Error reading {csv_path}: {e}")
        return None, None, None, None

def compute_resultant_vector(vectors):
    resultant_x = np.sum(vectors[:, 2])
    resultant_y = np.sum(vectors[:, 3])
    return np.array([resultant_x, resultant_y])

def compute_magnitude(vectors):
    return np.sqrt(vectors[:, 2]**2 + vectors[:, 3]**2)

def compute_angle(vectors):
    # Normalize angle range to [0, 360)
    # atan2 returns angle in radians, convert to degrees and normalize
    return (np.rad2deg(np.arctan2(vectors[:, 3], vectors[:, 2]))+360)%360

def get_batch_mode():
    app = ctk.CTk()
    app.title("Select Processing Mode")
    batch_mode = ctk.BooleanVar(value=False)

    def on_toggle():
        batch_mode.set(not batch_mode.get())

    toggle_btn = ctk.CTkButton(app, text="Toggle Batch Mode", command=on_toggle)
    toggle_btn.pack(pady=10)

    def on_confirm():
        app.quit()

    confirm_btn = ctk.CTkButton(app, text="Confirm", command=on_confirm)
    confirm_btn.pack(pady=10)

    app.mainloop()
    app.destroy()
    return batch_mode.get()


def batch_process(csv_files):
    """Process multiple CSV files maintaining exact variable names."""
    resultant_vectors = []
    average_magnitudes = []
    euclidean_distances = []
    path_lengths = []
    angles = []
    vectors = []

    for csv_file in natsorted(csv_files):
        try:
            vector, points, datetimes, reps = compute_vectors(csv_file)
            if vector is None or len(vector) == 0:
                print(f"Skipping {csv_file}: No valid data")
                continue

            # Ensure vector is 2D array
            vector = np.array(vector)
            if vector.ndim == 1:
                vector = vector.reshape(-1, 4)

            # Calculate metrics
            euclidean_distance = np.linalg.norm(vector[:, 2:4], axis=1)
            magnitudes = compute_magnitude(vector)
            angle = compute_angle(vector)
            resultant = compute_resultant_vector(vector)
            
            # Store results with filename
            file_name = os.path.basename(csv_file)
            resultant_vectors.append((file_name, resultant))
            average_magnitudes.append((file_name, np.mean(magnitudes)))
            euclidean_distances.append((file_name, euclidean_distance))
            path_length = np.sum(euclidean_distance)
            path_lengths.append((file_name, path_length))
            angles.append((file_name, angle))
            vectors.append((file_name, vector))

        except Exception as e:
            print(f"Error processing {csv_file}: {e}")
            continue

    if not vectors:
        print("No valid data processed")
        return None, None, None, None, None, None, None, None, None

    # Calculate means exactly as before
    mean_resultant = np.mean([r[1] for r in resultant_vectors], axis=0)
    mean_magnitude = np.mean([m[1] for m in average_magnitudes])
    mean_path_length = np.mean([p[1] for p in path_lengths])

    return (
        resultant_vectors,
        average_magnitudes,
        mean_resultant,
        mean_magnitude,
        euclidean_distances,
        mean_path_length,
        angles,
        vectors,
        path_lengths
    )    

def single_process(csv_path):
    if not csv_path:
        print("No file selected.")
        return
    vectors, points, datetimes, reps = compute_vectors(csv_path)
    magnitudes = compute_magnitude(vectors)
    angles = compute_angle(vectors)
    resultant = compute_resultant_vector(vectors)
    average_magnitude = np.mean(magnitudes)
    euclidean_distance = np.linalg.norm(vectors[:, 2:4], axis=1)
    path_length = np.sum(euclidean_distance)
    file_name = datetimes

    return (
        (file_name, vectors),
        (file_name, magnitudes),
        (file_name, angles),
        (file_name, resultant),
        (file_name, average_magnitude),
        (file_name, euclidean_distance),
        (file_name, path_length)
    )

def export_to_excel(batch_mode, csv_dir=None, csv_path=None):
    if batch_mode:
        if not csv_dir:
            print("No directory selected.")
            return
        os.chdir(csv_dir)
        csv_files = os.listdir(csv_dir)
        csv_files = [os.path.normpath(os.path.join(csv_dir, f)) for f in csv_files if f.endswith('.csv')]
        if not csv_files:
            print("No CSV files found in directory.")
            return
        (resultant_vectors, average_magnitudes, mean_resultant, mean_magnitude,
         euclidean_distances, mean_path_length, angles, vectors, path_lengths) = batch_process(csv_files=csv_files)
        print("Exporting batch results to Excel...")

        # Batch summary file
        summary_data = {
            'mean_resultant_dx': [mean_resultant[0]],
            'mean_resultant_dy': [mean_resultant[1]],
            'mean_magnitude': [mean_magnitude],
            'mean_path_length': [mean_path_length]
        }
        df_summary = pd.DataFrame(summary_data)
        df_path_lengths = pd.DataFrame(path_lengths, columns=['file', 'path_length'])

        with pd.ExcelWriter("batch_summary.xlsx", engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name="Summary", index=False, float_format="%.3f")
        print("Batch summary exported to batch_summary.xlsx")

        wb = openpyxl.load_workbook(output_filename_summary := "batch_summary.xlsx")
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 10
        wb.save(output_filename_summary)
        print(f"Exported {output_filename_summary}")

        # Per-file Excel export
        for i, (fname, vec) in enumerate(vectors):
            xy = vec[:, :2]
            dxdy = vec[:, 2:]
            df_points = pd.read_csv(os.path.join(csv_dir, fname), header=None)
            datetimes = df_points.iloc[:, 2] if df_points.shape[1] > 2 else None
            last_xy = df_points.iloc[[-1], :2].to_numpy()
            all_xy = np.vstack([xy, last_xy])
            all_dxdy = np.vstack([dxdy, [[np.nan, np.nan]]])
            df_vectors = pd.DataFrame(np.hstack([all_xy, all_dxdy]), columns=['x', 'y', 'dx', 'dy'])

            # Use datetimes and rep for the 'file' and 'rep' columns if available
            if datetimes is not None:
                dt_list = list(datetimes)
                reps = df_points.iloc[:, 3] if df_points.shape[1] > 3 else None
                rep_list = list(reps) if reps is not None else [np.nan] * len(dt_list)
                # For Vectors DataFrame (length N)
                expected_len = len(df_vectors)
                dt_list_vectors = dt_list[:expected_len-1] + [np.nan]
                rep_list_vectors = rep_list[:expected_len-1] + [np.nan]
                df_vectors.insert(0, 'file', dt_list_vectors)
                df_vectors.insert(1, 'rep', rep_list_vectors)
                # For per-vector DataFrames (length N-1, start from rep 1)
                perrow_file_col = dt_list[1:expected_len]
                perrow_rep_col = rep_list[1:expected_len]
            else:
                df_vectors.insert(0, 'file', fname)
                df_vectors.insert(1, 'rep', [np.nan] * len(df_vectors))
                perrow_file_col = [fname] * len(vec)
                perrow_rep_col = [np.nan] * len(vec)

            # Find corresponding stats
            mag = [m for f, m in average_magnitudes if f == fname][0]
            ang = [a for f, a in angles if f == fname][0]
            res = [r for f, r in resultant_vectors if f == fname][0]
            eucl = [e for f, e in euclidean_distances if f == fname][0]
            path = [p for f, p in path_lengths if f == fname][0]

            # Find all magnitudes for this file
            all_mags = compute_magnitude(vec)
            dt_list_short = perrow_file_col[:len(all_mags)]
            rep_list_short = perrow_rep_col[:len(all_mags)]
            df_magnitudes = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'magnitude': all_mags})
            df_angles = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'angle': ang})
            df_resultant = pd.DataFrame([{'file': fname, 'resultant_x': res[0], 'resultant_y': res[1]}])
            df_avg_mag = pd.DataFrame([{'file': fname, 'average_magnitude': mag}])
            df_path = pd.DataFrame([{'file': fname, 'path_length': path}])
            df_eucl = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'euclidean_distance': eucl})

                        # --- Plotting and image embedding for batch mode ---

            # Euclidean Distance KDE
            if len(df_eucl['euclidean_distance'].dropna()) > 1:
                ax_kde = df_eucl['euclidean_distance'].plot.kde()
                ax_kde.set_xlabel("Euclidean Distance")
                ax_kde.set_ylabel("Density")
                fig_kde = ax_kde.get_figure()
                img_data = io.BytesIO()
                fig_kde.savefig(img_data, format='png')
                plt.close(fig_kde)
                img_data.seek(0)
            else:
                img_data = None  # Or handle as needed

            # Euclidean Distance Histogram
            if len(df_eucl['euclidean_distance'].dropna()) > 1:
                hist_img_data = io.BytesIO()
                ax_hist = df_eucl['euclidean_distance'].plot.hist(bins=50, alpha=0.5)
                ax_hist.set_xlabel("Euclidean Distance")
                ax_hist.set_ylabel("Frequency")
                fig_hist = ax_hist.get_figure()
                fig_hist.savefig(hist_img_data, format='png')
                plt.close(fig_hist)
                hist_img_data.seek(0)
            else:
                hist_img_data = None

            # Angles KDE
            if len(df_angles['angle'].dropna()) > 1:
                kde_angle_img_data = io.BytesIO()
                ax_kde_angle = df_angles['angle'].plot.kde()
                ax_kde_angle.set_xlabel("Angle (degrees)")
                ax_kde_angle.set_ylabel("Density")
                fig_kde_angle = ax_kde_angle.get_figure()
                fig_kde_angle.savefig(kde_angle_img_data, format='png')
                plt.close(fig_kde_angle)
                kde_angle_img_data.seek(0)
            else:
                kde_angle_img_data = None

            # Angles Histogram
            if len(df_angles['angle'].dropna()) > 1:
                hist_angle_img_data = io.BytesIO()
                ax_hist_angle = df_angles['angle'].plot.hist(bins=50, alpha=0.5)
                ax_hist_angle.set_xlabel("Angle (degrees)")
                ax_hist_angle.set_ylabel("Frequency")
                fig_hist_angle = ax_hist_angle.get_figure()
                fig_hist_angle.savefig(hist_angle_img_data, format='png')
                plt.close(fig_hist_angle)
                hist_angle_img_data.seek(0)
            else:
                hist_angle_img_data = None


            output_filename = f"{os.path.splitext(fname)[0]}_results.xlsx"
            with pd.ExcelWriter(output_filename, engine='xlsxwriter') as writer:
                df_resultant.to_excel(writer, sheet_name="Resultant", index=False, float_format="%.3f")
                df_avg_mag.to_excel(writer, sheet_name="AverageMagnitude", index=False, float_format="%.3f")
                df_eucl.to_excel(writer, sheet_name="EuclideanDistances", index=False, float_format="%.3f")
                df_angles.to_excel(writer, sheet_name="Angles", index=False, float_format="%.3f")
                df_vectors.to_excel(writer, sheet_name="Vectors", index=False, float_format="%.3f")
                df_magnitudes.to_excel(writer, sheet_name="Magnitudes", index=False, float_format="%.3f")
                df_path.to_excel(writer, sheet_name="PathLengths", index=False, float_format="%.3f")

                workbook = writer.book
                worksheet = writer.sheets['EuclideanDistances']
                worksheet_angles = writer.sheets['Angles']

                if img_data is not None:
                    worksheet.insert_image('F5', 'eucl_kde.png', {'image_data': img_data})
                if hist_img_data is not None:
                    worksheet.insert_image('F30', 'eucl_hist.png', {'image_data': hist_img_data})
                if kde_angle_img_data is not None:
                    worksheet_angles.insert_image('F2', 'angle_kde.png', {'image_data': kde_angle_img_data})
                if hist_angle_img_data is not None:
                    worksheet_angles.insert_image('F30', 'angle_hist.png', {'image_data': hist_angle_img_data})

            # Auto-fit columns for all sheets
            wb = openpyxl.load_workbook(output_filename)
            for ws in wb.worksheets:
                for col in ws.columns:
                    max_length = 0
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            cell_value = str(cell.value)
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 10
            wb.save(output_filename)
            print(f"Exported {output_filename}")

    else:
        
        if not csv_path:
            print("No file selected.")
            return
        os.chdir(os.path.dirname(csv_path))
        single_results = single_process(csv_path=csv_path)
        if single_results is None:
            print("No results to export.")
            return
        vectors, magnitudes, angles, resultant, average_magnitude, euclidean_distance, path_length = single_results
        vectors = [vectors]
        magnitudes = [magnitudes]
        angles = [angles]
        resultant_vectors = [resultant]
        average_magnitudes = [average_magnitude]
        euclidean_distances = [euclidean_distance]
        path_lengths = [path_length]

        # Get datetimes from the CSV
        df_points = pd.read_csv(csv_path, header=None)
        datetimes = df_points.iloc[:, 2] if df_points.shape[1] > 2 else None

        # Vectors: shape (N, 4)
        xy = vectors[0][1][:, :2]
        dxdy = vectors[0][1][:, 2:]

        # Get last x, y from original CSV
        last_xy = df_points.iloc[[-1], :2].to_numpy()

        # Stack x, y and dx, dy
        all_xy = np.vstack([xy, last_xy])
        all_dxdy = np.vstack([dxdy, [[np.nan, np.nan]]])

        df_vectors = pd.DataFrame(
            np.hstack([all_xy, all_dxdy]),
            columns=['x', 'y', 'dx', 'dy']
        )

        if datetimes is not None:
            dt_list = list(datetimes)
            reps = df_points.iloc[:, 3] if df_points.shape[1] > 3 else None
            rep_list = list(reps) if reps is not None else [np.nan] * len(dt_list)
            expected_len = len(df_vectors)
            # For Vectors DataFrame (length N)
            dt_list_vectors = dt_list[:expected_len-1] + [np.nan]
            rep_list_vectors = rep_list[:expected_len-1] + [np.nan]
            df_vectors.insert(0, 'file', dt_list_vectors)
            df_vectors.insert(1, 'rep', rep_list_vectors)
            # For per-vector DataFrames (length N-1, start from rep 1)
            perrow_file_col = dt_list[1:expected_len]
            perrow_rep_col = rep_list[1:expected_len]
            dt_list_short = perrow_file_col[:len(magnitudes[0][1])]
            rep_list_short = perrow_rep_col[:len(magnitudes[0][1])]
            df_magnitudes = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'magnitude': magnitudes[0][1]})
            df_angles = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'angle': angles[0][1]})
            df_eucl = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'euclidean_distance': euclidean_distance[1]})
        else:
            df_vectors.insert(0, 'file', vectors[0][0])
            df_vectors.insert(1, 'rep', [np.nan] * len(df_vectors))
            perrow_file_col = [vectors[0][0]] * len(vectors[0][1])
            perrow_rep_col = [np.nan] * len(vectors[0][1])
            dt_list_short = perrow_file_col[:len(magnitudes[0][1])]
            rep_list_short = perrow_rep_col[:len(magnitudes[0][1])]
            df_magnitudes = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'magnitude': magnitudes[0][1]})
            df_angles = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'angle': angles[0][1]})
            df_eucl = pd.DataFrame({'file': dt_list_short, 'rep': rep_list_short, 'euclidean_distance': euclidean_distance[1]})

        # Resultant: shape (2,)
        filename = os.path.basename(csv_path)
        df_resultant = pd.DataFrame([{'file': filename, 'resultant_x': resultant[1][0], 'resultant_y': resultant[1][1]}])
        # Average magnitude: scalar
        df_avg_mag = pd.DataFrame([{'file': filename, 'average_magnitude': average_magnitude[1]}])
        # Path length: scalar
        df_path = pd.DataFrame([{'file': filename, 'path_length': path_length[1]}])

        # Ensure 'rep' is numeric for all relevant DataFrames before writing to Excel
        for df in [df_eucl, df_angles, df_magnitudes, df_vectors]:
            if 'rep' in df.columns:
                df['rep'] = pd.to_numeric(df['rep'], errors='coerce')

        output_filename = f"{os.path.splitext(os.path.basename(csv_path))[0]}_results.xlsx"

       # Euclidean Distance KDE
        if len(df_eucl['euclidean_distance'].dropna()) > 1:
            ax_kde = df_eucl['euclidean_distance'].plot.kde()
            ax_kde.set_xlabel("Euclidean Distance")
            ax_kde.set_ylabel("Density")
            fig_kde = ax_kde.get_figure()
            img_data = io.BytesIO()
            fig_kde.savefig(img_data, format='png')
            plt.close(fig_kde)
            img_data.seek(0)
        else:
            img_data = None  # Or handle as needed

        # Euclidean Distance Histogram
        if len(df_eucl['euclidean_distance'].dropna()) > 1:
            hist_img_data = io.BytesIO()
            ax_hist = df_eucl['euclidean_distance'].plot.hist(bins=50, alpha=0.5)
            ax_hist.set_xlabel("Euclidean Distance")
            ax_hist.set_ylabel("Frequency")
            fig_hist = ax_hist.get_figure()
            fig_hist.savefig(hist_img_data, format='png')
            plt.close(fig_hist)
            hist_img_data.seek(0)
        else:
            hist_img_data = None

        # Angles KDE
        if len(df_angles['angle'].dropna()) > 1:
            kde_angle_img_data = io.BytesIO()
            ax_kde_angle = df_angles['angle'].plot.kde()
            ax_kde_angle.set_xlabel("Angle (degrees)")
            ax_kde_angle.set_ylabel("Density")
            fig_kde_angle = ax_kde_angle.get_figure()
            fig_kde_angle.savefig(kde_angle_img_data, format='png')
            plt.close(fig_kde_angle)
            kde_angle_img_data.seek(0)
        else:
            kde_angle_img_data = None

        # Angles Histogram
        if len(df_angles['angle'].dropna()) > 1:
            hist_angle_img_data = io.BytesIO()
            ax_hist_angle = df_angles['angle'].plot.hist(bins=50, alpha=0.5)
            ax_hist_angle.set_xlabel("Angle (degrees)")
            ax_hist_angle.set_ylabel("Frequency")
            fig_hist_angle = ax_hist_angle.get_figure()
            fig_hist_angle.savefig(hist_angle_img_data, format='png')
            plt.close(fig_hist_angle)
            hist_angle_img_data.seek(0)
        else:
            hist_angle_img_data = None

        with pd.ExcelWriter(output_filename, engine='xlsxwriter') as writer:
            df_resultant.to_excel(writer, sheet_name="Resultant", index=False, float_format="%.3f")
            df_avg_mag.to_excel(writer, sheet_name="AverageMagnitude", index=False, float_format="%.3f")
            df_eucl.to_excel(writer, sheet_name="EuclideanDistances", index=False, float_format="%.3f")
            df_angles.to_excel(writer, sheet_name="Angles", index=False, float_format="%.3f")
            df_vectors.to_excel(writer, sheet_name="Vectors", index=False, float_format="%.3f")
            df_magnitudes.to_excel(writer, sheet_name="Magnitudes", index=False, float_format="%.3f")
            df_path.to_excel(writer, sheet_name="PathLengths", index=False, float_format="%.3f")

            workbook = writer.book
            worksheet = writer.sheets['EuclideanDistances']
            worksheet_angles = writer.sheets['Angles']

            if img_data is not None:
                worksheet.insert_image('F5', 'eucl_kde.png', {'image_data': img_data})
            if hist_img_data is not None:
                worksheet.insert_image('F30', 'eucl_hist.png', {'image_data': hist_img_data})
            if kde_angle_img_data is not None:
                worksheet_angles.insert_image('F2', 'angle_kde.png', {'image_data': kde_angle_img_data})
            if hist_angle_img_data is not None:
                worksheet_angles.insert_image('F30', 'angle_hist.png', {'image_data': hist_angle_img_data})

        # Auto-fit columns for all sheets (do this AFTER closing the ExcelWriter)
        wb = openpyxl.load_workbook(output_filename)
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 10
        wb.save(output_filename)

        print("Single file results exported to", output_filename)

def colnum_string(n):
    """Convert zero-based column index to Excel column letter."""
    string_ = ""
    while n >= 0:
        string_ = chr(n % 26 + ord('A')) + string_
        n = n // 26 - 1
    return string_

if __name__ == "__main__":
    pass